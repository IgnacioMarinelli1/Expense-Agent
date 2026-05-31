import csv
import io
import logging

logger = logging.getLogger(__name__)

SUPPORTED_SPREADSHEET_MIME = {
    "text/csv",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}

_MAX_ROWS = 300


def _csv_bytes_to_text(data: bytes, filename: str) -> str:
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return f"Archivo vacío: {filename}"
    lines = [f"Archivo: {filename}"]
    for i, row in enumerate(rows[:_MAX_ROWS]):
        lines.append(" | ".join(str(c) for c in row))
        if i == 0 and len(rows) > 1:
            lines.append("|".join(["---"] * len(row)))
    if len(rows) > _MAX_ROWS:
        lines.append(f"... ({len(rows)} filas en total, se muestran las primeras {_MAX_ROWS})")
    return "\n".join(lines)


def _xlsx_bytes_to_text(data: bytes, filename: str) -> str:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl no instalado — ejecutá: pip install openpyxl")
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines = [f"Archivo: {filename}"]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"\nHoja: {sheet_name}")
        count = 0
        first_row = True
        for row in ws.iter_rows(values_only=True):
            if all(c is None for c in row):
                continue
            if count >= _MAX_ROWS:
                lines.append(f"... (se omiten filas adicionales, límite {_MAX_ROWS})")
                break
            lines.append(" | ".join(str(c) if c is not None else "" for c in row))
            if first_row:
                lines.append("|".join(["---"] * len(row)))
                first_row = False
            count += 1
    wb.close()
    return "\n".join(lines)


def spreadsheet_to_text(data: bytes, mime_type: str, filename: str = "archivo") -> str:
    if mime_type == "text/csv":
        return _csv_bytes_to_text(data, filename)
    return _xlsx_bytes_to_text(data, filename)
